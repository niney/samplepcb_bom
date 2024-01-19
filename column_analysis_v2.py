import collections
import json
import os
from difflib import SequenceMatcher

import numpy as np
import openpyxl
import pandas as pd
import requests as reqs
import xlrd
from openpyxl.workbook import Workbook
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.linear_model import LogisticRegression
from sklearn.metrics import accuracy_score
from sklearn.model_selection import train_test_split

from data_extractor import DataExtractor
from direct_analysis import DirctAnalysis
from parts_analysis import PartsAnalysis
from parts_export_service import PartsExportService
from samplepcb_logger import make_logger
from search_service import SearchService

extractor = DataExtractor()
parts_analyzer = PartsAnalysis()
direct_analysis = DirctAnalysis(parts_analyzer, extractor)
search_service = SearchService()
parts_export_service = PartsExportService()

tfidf_vect = TfidfVectorizer(stop_words='english', ngram_range=(1, 1), max_df=100)
lr_clf = LogisticRegression(C=5)

bom_item_texts = []
bom_item_label = []

REFERENCE = 1
PART_NUMBER = 2
QTY = 4
PACKAGE = 6
VALUE = 9
OTHER = 15

NO = 99
EXCEL_CAL = 100
URL = 98

logger = make_logger('analysis_v2')


# 엑셀 파일 로드
# 컬럼 카운트 리스트, sheet 반환
def load_bom_excel(filename, sheet_idx=0):
    file_path = './upload/' + filename
    fname, ext = os.path.splitext(filename)
    if ext == '.xls':
        cvt_xls_to_xlsx(file_path, file_path + 'x')
        file_path = file_path + 'x'

    bom_excel = openpyxl.load_workbook(file_path, data_only=True)

    # 행별 실제 사용하는 컬럼 개수 알아내기
    sheet = bom_excel.worksheets[sheet_idx]

    column_cnt_list = {}
    for idx, row in enumerate(sheet.rows):
        none_cnt = 0
        for r in row:
            if r.value is not None:
                none_cnt = none_cnt + 1
            column_cnt_list[idx] = none_cnt

    return column_cnt_list, sheet


# https://stackoverflow.com/questions/9918646/how-to-convert-xls-to-xlsx/42574983#42574983
def cvt_xls_to_xlsx(src_file_path, dst_file_path):
    book_xls = xlrd.open_workbook(src_file_path)
    book_xlsx = Workbook()

    sheet_names = book_xls.sheet_names()
    for sheet_index, sheet_name in enumerate(sheet_names):
        sheet_xls = book_xls.sheet_by_name(sheet_name)
        if sheet_index == 0:
            sheet_xlsx = book_xlsx.active
            sheet_xlsx.title = sheet_name
        else:
            sheet_xlsx = book_xlsx.create_sheet(title=sheet_name)

        for row in range(0, sheet_xls.nrows):
            for col in range(0, sheet_xls.ncols):
                sheet_xlsx.cell(row=row + 1, column=col + 1).value = sheet_xls.cell_value(row, col)

    book_xlsx.save(dst_file_path)


# 가장 큰 값 알아내기
def get_max_count(tuple_data):
    data_count = collections.Counter(val for val in tuple_data.values())

    def f1(x):
        return data_count[x]

    max_cnt = max(data_count.keys(), key=f1)
    return max_cnt


# tuple 데이터로부터 최대값을 가진 인덱스를 반환
def get_index_max_value(tuple_data):
    def f1(x):
        return tuple_data[x]

    return max(tuple_data, key=f1)


def search_pcb_header_each(sheet):
    sentence_results = []
    score_tuple = {}
    score_tuple_for_bom = {}
    for (idx, rows) in enumerate(sheet.rows):
        row_list = []
        row_list_for_bom = []
        for row in rows:
            if row.value is not None:
                row_list.append(str(row.value))
            else:
                row_list.append('')

            column_target = None
            if row.value is None:
                column_target = None
            row_list_for_bom.append({'type': type(row.value).__name__, 'columnIdx': idx, 'columnTarget': column_target,
                                     'query': str(row.value).strip()})

        URL = 'http://localhost:8080/api/pcbColumn/_searchSentenceList'
        response = reqs.post(URL, json={'queryColumnNameList': row_list})
        body = json.loads(response.text)
        response_data = body['data']
        score_tuple[idx] = response_data['averageScore']
        sentence_results.append(response_data)
        score_tuple_for_bom[idx] = bom_ml(row_list_for_bom, True)
        if idx == 13:
            break

    header_column_idx = get_index_max_value(score_tuple)
    if score_tuple_for_bom[header_column_idx]['averageScore'] < 50:
        header_column_idx = header_column_idx + 1
    return {'headerColumnIdx': header_column_idx + 1, 'headerDetail': sentence_results[header_column_idx]}


def bom_ml_test():
    samplepcb_df = pd.read_csv('./data/samplepcb_bom.csv', header=None)

    X = samplepcb_df.iloc[:, :-1]
    y = samplepcb_df.iloc[:, -1]

    X_train, X_test, y_train, y_test = train_test_split(X, y, test_size=0.2, random_state=156, stratify=y)

    X_train = X_train[0].tolist()
    X_test = X_test[0].tolist()
    tfidf_vect = TfidfVectorizer(stop_words='english', ngram_range=(1, 1), max_df=100)
    tfidf_vect.fit(X_train)
    X_train_tfidf_vect = tfidf_vect.transform(X_train)
    X_test_tfidf_vect = tfidf_vect.transform(X_test)

    lr_clf = LogisticRegression(C=5)
    lr_clf.fit(X_train_tfidf_vect, y_train)
    pred = lr_clf.predict(X_test_tfidf_vect)
    print('TF-IDF Vectorized Logistic Regression 의 예측 정확도는 {0:.3f}'.format(accuracy_score(y_test, pred)))


def bom_ml_tuning():
    samplepcb_df = pd.read_csv('./data/samplepcb_bom.csv', header=None)

    X = samplepcb_df.iloc[:, :-1]
    y = samplepcb_df.iloc[:, -1]

    X_train, X_test, y_train, y_test = train_test_split(X, y, test_size=0.2, random_state=156, stratify=y)

    X_train = X_train[0].tolist()
    X_test = X_test[0].tolist()

    from sklearn.model_selection import GridSearchCV
    from sklearn.pipeline import Pipeline

    pipeline = Pipeline([
        ('tfidf_vect', TfidfVectorizer(stop_words='english')),
        ('lr_clf', LogisticRegression())
    ])

    # Pipeline에 기술된 각각의 객체 변수에 언더바(_)2개를 연달아 붙여 GridSearchCV에 사용될
    # 파라미터/하이퍼 파라미터 이름과 값을 설정. .
    params = {'tfidf_vect__ngram_range': [(1, 1), (1, 2), (1, 3)],
              'tfidf_vect__max_df': [100, 300, 700],
              'lr_clf__C': [1, 5, 10]
              }

    # GridSearchCV의 생성자에 Estimator가 아닌 Pipeline 객체 입력
    grid_cv_pipe = GridSearchCV(pipeline, param_grid=params, cv=3, scoring='accuracy', verbose=1)
    grid_cv_pipe.fit(X_train, y_train)
    print(grid_cv_pipe.best_params_, grid_cv_pipe.best_score_)

    pred = grid_cv_pipe.predict(X_test)
    print('Pipeline을 통한 Logistic Regression 의 예측 정확도는 {0:.3f}'.format(accuracy_score(y_test, pred)))


def bom_ml_init():
    global tfidf_vect
    global lr_clf

    samplepcb_df = pd.concat(pd.read_excel('./data/samplepcb_bom.xlsx', header=None, sheet_name=None),
                             ignore_index=True)

    X = samplepcb_df.iloc[:, :-1]
    y = samplepcb_df.iloc[:, -1]

    X_train = [str(x) for x in X[0].tolist()]
    y_train = y.tolist()

    tfidf_vect.fit(X_train)
    X_train_tfidf_vect = tfidf_vect.transform(X_train)

    lr_clf.fit(X_train_tfidf_vect, y_train)


def bom_ml_init_by_api():
    global tfidf_vect
    global lr_clf

    tfidf_vect = TfidfVectorizer(stop_words='english', ngram_range=(1, 1), max_df=100)
    lr_clf = LogisticRegression(C=5)

    URL = 'http://localhost:8080/api/pcbItem/_allItemGroupByTarget'
    response = reqs.get(URL)
    body = json.loads(response.text)

    X_train = []
    y_train = []
    for items in body['data']:
        for item in items:
            X_train.append(item['itemName'])
            y_train.append(item['target'])

    tfidf_vect.fit(X_train)
    X_train_tfidf_vect = tfidf_vect.transform(X_train)

    lr_clf.fit(X_train_tfidf_vect, y_train)


def bom_ml(query_list, zero_score=False):
    global tfidf_vect
    global lr_clf

    query_values = [l['query'] for l in query_list]
    query_tfidf_vect = tfidf_vect.transform(query_values)
    predict_list = lr_clf.predict(query_tfidf_vect)
    predict_proba_list = lr_clf.predict_proba(query_tfidf_vect)
    score_list = []
    results = []
    for idx, meta in enumerate(query_list):
        score = max(predict_proba_list[idx]) * 100
        if type(None).__name__ == meta['type']:
            score = 0
            if zero_score:
                score_list.append(score)
        else:
            score_list.append(score)

        meta['predict'] = predict_list[idx].item()
        meta['score'] = score
        # classification 결과를 받아옴
        classification_result = {k: v for k, v in classification(meta['query']).items() if k != 'productName'}
        # classification_result가 비어 있지 않은 경우에만 meta['classification']에 값을 넣음
        if classification_result:
            meta['classification'] = classification_result
        results.append(meta)

    if len(score_list) == 0:
        score_list.append(0)

    return {'predictResults': results, 'averageScore': np.mean(score_list)}


def trim_none_or_empty_from_end(strings):
    # 리스트 끝에서부터 None, 'None' 또는 빈 문자열이 아닌 첫 번째 요소 찾기
    first_non_none_or_empty = next((i for i, item in enumerate(reversed(strings)) if item and item != "None"),
                                   len(strings))

    # 해당 인덱스를 사용하여 필요한 부분만 남기기
    return strings[:len(strings) - first_non_none_or_empty]


def classification(item_val, reference_prefix=None):
    title_text = parts_analyzer.split_text(item_val)
    return parts_analyzer.parse_string(title_text, reference_prefix)


def find_highest_score_per_target(data):
    highest_scores = {}

    for item in data:
        target = item['target']
        score = item['_score']

        # target이 이미 highest_scores에 있고 현재 score가 더 높은 경우에만 업데이트
        if target not in highest_scores or score > highest_scores[target]['_score']:
            highest_scores[target] = item

    # 각 target별로 가장 높은 점수를 받은 항목을 리스트로 반환
    return list(highest_scores.values())


def bom_ml_cols(query_list):
    global tfidf_vect
    global lr_clf

    query_values = [l['query'] for l in query_list]
    query_tfidf_vect = tfidf_vect.transform(query_values)
    predict_list = lr_clf.predict(query_tfidf_vect)
    predict_proba_list = lr_clf.predict_proba(query_tfidf_vect)
    score_list = []
    results = []
    for idx, meta in enumerate(query_list):
        score = max(predict_proba_list[idx]) * 100
        if type(None).__name__ == meta['type']:
            score = 0
            # score_list.append(score) # None, 0 일때는 평균점수에 반영하지 않는다
        if meta['query'] == 'None':
            score = 0
        else:
            score_list.append(score)

        if len(score_list) == 0:
            score_list.append(0)

        meta['predict'] = predict_list[idx].item()
        meta['score'] = score
        results.append(meta)

    queries = [result['query'] for result in results]
    # 리스트 끝에서부터 None, 'None' 또는 빈 문자열이 아닌 첫 번째 요소 찾기
    queries = trim_none_or_empty_from_end(queries)
    if direct_analysis.percent_reference(queries, 55) >= 55:
        return {'predict': REFERENCE, 'predictResults': results, 'averageScore': 100}
    if direct_analysis.percent_classification(queries) >= 40:
        return {'predict': VALUE, 'predictResults': results, 'averageScore': 100}
    if direct_analysis.percent_package(queries) >= 60:
        return {'predict': PACKAGE, 'predictResults': results, 'averageScore': 100}
    if direct_analysis.percent_increment_digit_by_list(queries) >= 70:
        return {'predict': 99, 'predictResults': results, 'averageScore': 100}  # 99 는 No
    if direct_analysis.percent_digit_by_list(queries) >= 75:
        return {'predict': 4, 'predictResults': results, 'averageScore': 100}  # 4 는 수량
    if direct_analysis.starts_with_by_list(queries, '=') >= 75:
        return {'predict': 100, 'predictResults': results, 'averageScore': 100}  # 100 는 엑셀서식
    if direct_analysis.percent_valid_urls(queries) >= 75:
        return {'predict': 98, 'predictResults': results, 'averageScore': 100}  # 98 는 URL

    predicts = [result['predict'] for result in results]

    data_count = collections.Counter(predicts)

    def f1(x):
        return data_count[x]

    max_cnt = max(data_count.keys(), key=f1)

    # 예상된 분류의 점수값만 추가하여 평균계산
    predict_score_list = []
    none_score = 0
    for result in results:
        if result['predict'] == max_cnt and result['score'] != 0:
            predict_score_list.append(result['score'])
        if result['score'] == 0:
            none_score = none_score + 1

    if len(predict_score_list) == 0:
        predict_score_list.append(0)

    minus_score = (none_score / len(results)) * 100
    return {'predict': max_cnt, 'predictResults': results, 'averageScore': np.mean(predict_score_list) - minus_score}


def search_pcb_column_each(sheet, start_item_index, header_column_search_list):
    """
PCB 컬럼 분석검색
    :param sheet: 시트
    :param start_item_index: item 시작 인덱스
    :param header_column_search_list: 분석된 헤더컬럼 정보
    :return:
    """
    ml_results = []
    column_idx = start_item_index
    for col_idx, rows in enumerate(sheet.iter_rows(min_row=start_item_index)):
        row_list = []
        ml_result = {}
        for idx, row in enumerate(rows):
            column_target = None
            if len(header_column_search_list) > idx:
                column_target = header_column_search_list[idx]['target']
            if row.value is None:
                column_target = None
            row_list.append({'type': type(row.value).__name__, 'columnIdx': idx, 'columnTarget': column_target,
                             'query': str(row.value).strip()})

        is_pcb_item = 'isPcbItem'
        if len(row_list) != 0:
            if direct_analysis.percent_none_for_is_pcb([l['query'] for l in row_list], 90):
                ml_result[is_pcb_item] = False
            else:
                ml_result = bom_ml(row_list)
                if ml_result['averageScore'] > 50:
                    ml_result[is_pcb_item] = True
                else:
                    ml_result[is_pcb_item] = False
        else:
            ml_result[is_pcb_item] = False

        ml_result['columnIdx'] = column_idx
        column_idx = column_idx + 1
        ml_results.append(ml_result)

    return ml_results


def search_pcb_column_cols(sheet, start_item_index, header_column_search_list):
    ml_results = []
    for col_idx, cols in enumerate(sheet.iter_cols(min_row=start_item_index)):
        col_list = []
        ml_result = {}
        for row_idx, col in enumerate(cols):
            column_target = None
            if len(header_column_search_list) > row_idx:
                column_target = header_column_search_list[col_idx]['target']
            if col.value is None:
                column_target = None
            col_list.append({'type': type(col.value).__name__, 'rowIdx': row_idx, 'columnTarget': column_target,
                             'query': str(col.value).strip()})

        if len(col_list) != 0:
            ml_result = bom_ml_cols(col_list)

        ml_results.append(ml_result)

    search_service.find_update_part_number_average_score(ml_results)
    return ml_results


def search_pcb_header_each_pandas(df):
    """
    pandas DataFrame에서 헤더 열을 API 요청을 통해 검색하고 ML 모델을 호출하여 찾습니다.

    Args:
        df (pandas.DataFrame): 검색할 DataFrame입니다.

    Returns:
        dict: 헤더 열의 인덱스와 헤더에 대한 자세한 정보를 포함하는 딕셔너리입니다.
            - 'headerColumnIdx': 헤더 열의 인덱스입니다.
            - 'headerDetail': 헤더 열에 대한 자세한 정보입니다.
    """
    sentence_results = []
    score_tuple = {}
    score_tuple_for_bom = {}

    for idx, row in df.iterrows():
        row_list = row.fillna('').astype(str).tolist()
        row_list_for_bom = [
            {'type': type(val).__name__,
             'columnIdx': idx,
             'columnTarget': None,
             'query': str(val).strip()} for val in row
        ]

        # 실제 API에 기반하여 적절하게 수정해야 하는 API 요청 예시입니다.
        URL = 'http://localhost:8080/api/pcbColumn/_searchSentenceList'
        response = reqs.post(URL, json={'queryColumnNameList': row_list})
        body = json.loads(response.text)
        response_data = body['data']
        score_tuple[idx] = response_data['averageScore']
        sentence_results.append(response_data)

        score_tuple_for_bom[idx] = search_service.check_is_pcb_item(row_list_for_bom)

        # Break condition - this might need to be adapted
        if idx == 13:
            break

    # Logic to determine the header column index
    header_column_idx = get_index_max_value(score_tuple)
    header_column_real_idx = header_column_idx
    if (header_column_real_idx + 1 in score_tuple_for_bom
            and not score_tuple_for_bom[header_column_real_idx + 1]['isPcbItem']):
        header_column_real_idx = header_column_real_idx + 1

    return {'headerColumnIdx': header_column_real_idx, 'headerDetail': sentence_results[header_column_idx]}


def search_pcb_column_each_pandas(df, start_item_index, header_column_search_list):
    ml_results = []
    # start_item_index = 14
    column_idx = start_item_index

    for row_idx, row in df.iterrows():
        # 행 인덱스가 start_item_index 미만인 경우 건너뜁니다.
        if row_idx < start_item_index:
            continue

        row_list = []
        ml_result = {}

        # 행의 각 셀에 대해 처리합니다.
        gl_score = 0
        for col_idx, val in enumerate(row):
            target = None
            if len(header_column_search_list) > col_idx:
                target = header_column_search_list[col_idx].get('target', None)
                gl_score = header_column_search_list[col_idx].get('glScore', None)

            row_list.append({
                'type': type(val).__name__,
                'columnIdx': col_idx,
                'target': target,
                'query': str(val).strip() if pd.notnull(val) else '',
                'glScore': gl_score
            })

        # target 업데이트
        set_best_target(row_list)

        analysis_result = all_analyze_bom(row_list)
        # all_analyze_bom_ml(analysis_result)
        analysis_result['columnIdx'] = row_idx
        analysis_result['rowList'] = row_list
        extract_result = parts_export_service.extract_values_by_analysis_result(analysis_result['classificationResult'])
        if not parts_analyzer.is_float(extract_result):
            analysis_result['classificationResult']['searchText'] = extract_result
        ml_results.append(analysis_result)

    return ml_results


def set_best_target(bom_list):
    # target 값에 따른 가장 높은 glScore를 추적하기 위한 딕셔너리 초기화
    highest_scores = {}

    # 데이터를 반복하며 highest_scores를 업데이트
    for item in bom_list:
        target = item['target']
        gl_socre = item['glScore']

        if gl_socre is None:
            continue

        # 현재 target에 대해 가장 높은 glScore를 추적
        if target not in highest_scores or gl_socre > highest_scores[target]['glScore']:
            highest_scores[target] = {'glScore': gl_socre, 'target': target}

    # 모든 항목에 대해 bestTarget 필드 추가
    for item in bom_list:
        target = item['target']

        if target is None:
            continue
        # 현재 항목의 target이 해당 target의 최고 glScore를 가진 항목의 target과 같으면 bestTarget으로 설정
        if item['glScore'] == highest_scores[target]['glScore']:
            item['target'] = target
        else:
            item['target'] = -1


def all_analyze_bom(bom_item_list):
    classification_result = {}
    reference_pattern = None
    other = []
    filtered_bom_item_list = []
    for item in bom_item_list:
        if item['target'] == REFERENCE:
            classification_result['reference'] = item['query']
            reference_pattern = direct_analysis.find_reference_pattern(item['query'])
            if reference_pattern is not None:
                if reference_pattern == 'C':
                    classification_result['pcbType'] = 'capacitor'
                elif reference_pattern == 'R':
                    classification_result['pcbType'] = 'resistor'
        elif item['target'] == QTY:
            classification_result['qty'] = item['query']
        elif item['target'] == NO:
            classification_result['no'] = item['query']
        elif item['target'] == EXCEL_CAL:
            classification_result['excelCal'] = item['query']
        elif item['target'] == URL:
            classification_result['url'] = item['query']
        elif item['target'] == PACKAGE and 'glScore' in item and item['glScore'] > 99:
            classification_result['package'] = item['query']
        elif item['target'] == PART_NUMBER:
            part_classify_result = classification(item['query'], reference_pattern)
            # 유효한 부품 번호만을 포함할 새로운 리스트 생성
            valid_part_numbers = []
            for part_number in part_classify_result['productName']:
                if parts_analyzer.is_part_number(part_number):
                    valid_part_numbers.append(part_number)
                else:
                    other.append(part_number)

            if valid_part_numbers:
                # 각 부품 번호의 점수를 받음
                part_number_scores = search_service.find_pcb_item_score(PART_NUMBER, valid_part_numbers)

                # 점수가 50점이상 부품 번호만 유지하는 리스트
                valid_part_numbers = [part_number for part_number in valid_part_numbers if
                                      part_number_scores.get(part_number) > 50]

            # 새로운 유효한 부품 번호 리스트로 업데이트
            classification_result['partNumber'] = valid_part_numbers
            filtered_bom_item_list.append(item)
        else:
            filtered_bom_item_list.append(item)

    excluded_targets = [REFERENCE, QTY, NO, EXCEL_CAL, URL]
    update_query = parts_analyzer.update_component_queries(bom_item_list, excluded_targets)
    if update_query:
        other = [element for element in other if element != update_query]

    bom_classify_result = parts_analyzer.analyze_and_classify_bom(filtered_bom_item_list, reference_pattern)
    original_classification = bom_classify_result['classificationResult']['originalClassification']
    original_classification.update(classification_result)

    source_part_numbers = []
    product_names = original_classification['productName']
    for part_number in product_names:
        if parts_analyzer.is_part_number(part_number):
            source_part_numbers.append(part_number)
        else:
            other.append(part_number)

    if source_part_numbers:
        if 'partNumber' not in original_classification:
            original_classification['partNumber'] = []
        # octopart에서 부품 번호를 찾음
        part_number_scores = search_service.find_pcb_item_score(PART_NUMBER, source_part_numbers)
        find_part_numbers = [part for part in part_number_scores if
                             part_number_scores.get(part) == 100]
        if find_part_numbers:
            original_classification['partNumber'].append(find_part_numbers[0])
        else:
            other_scores = search_service.find_pcb_item_score(OTHER, source_part_numbers)
            other = other + [other for other in other_scores if  # 검색에 있다면 합치기
                             other_scores.get(other) == 100]

            source_part_numbers = [other for other in other_scores if
                                   other_scores.get(other) != 100]
            for source_part_number in source_part_numbers:
                search_result = search_service.search_octopart_mpn(source_part_number)
                if 'same' in search_result and search_result['same']:
                    original_classification['partNumber'].append(source_part_number)
                    indexing_pcb_item(PART_NUMBER, source_part_number)
                    break
                elif 'like' in search_result and search_result['like']:
                    if 'candidatePartNumber' not in original_classification:
                        original_classification['candidatePartNumber'] = []
                        if not parts_analyzer.is_float(source_part_number): # 숫자형이 아니어야 후보 부품 번호로 분류
                            original_classification['candidatePartNumber'].append(source_part_number)
                else:
                    other.append(source_part_number)
                    indexing_pcb_item(OTHER, source_part_number)

    # 중복제거
    if 'partNumber' in original_classification:
        original_classification['partNumber'] = list(set(original_classification['partNumber']))
        source_part_numbers = \
            [item for item in source_part_numbers if item not in set(original_classification['partNumber'])]
    if 'candidatePartNumber' in original_classification:
        source_part_numbers = \
            [item for item in source_part_numbers if item not in set(original_classification['candidatePartNumber'])]

    original_classification['productName'] = source_part_numbers
    original_classification['other'] = other

    return bom_classify_result


def all_analyze_bom_ml(bom_classify_result):
    original_classification = bom_classify_result['classificationResult']['originalClassification']
    if (('partNumber' not in original_classification or not original_classification['partNumber'])
            and
            ('candidatePartNumber' not in original_classification or not original_classification['candidatePartNumber'])
            and
            ('productName' not in original_classification or not original_classification['productName'])
            and
            ('flatClassification' not in bom_classify_result['classificationResult'] or
             ('flatClassification' in bom_classify_result['classificationResult'] and
              not bom_classify_result['classificationResult']['flatClassification']))):
        origin_text = bom_classify_result['parsedData']['originText']
        if origin_text is None or not origin_text.strip():
            return

        reference_prefix = None
        if 'pcbType' in original_classification:
            if original_classification['pcbType'] == 'capacitor':
                reference_prefix = 'C'
            elif original_classification['pcbType'] == 'resistor':
                reference_prefix = 'R'

        # 한번 더 분석
        cst = parts_analyzer.custom_split_texts(bom_classify_result['parsedData']['words'])
        cst_parsed = parts_analyzer.parse_string(cst, reference_prefix)
        parsed_value_llama_dict = {k: v for k, v in cst_parsed.items() if k not in "productName"}
        if not bool(parsed_value_llama_dict):
            response_json = search_service.req_llama(origin_text)
            original_classification['candidateText'] = response_json['response']
            llama_dict = parts_export_service.parse_key_value_by_llama(response_json['response'])
            extract_value_llama_dict = parts_export_service.extract_values_from_parsed_data(
                llama_dict, ["BrandName", "Manufacturer", "Vendor", "Unit"])
            logger.info(origin_text + ' : ' + str(extract_value_llama_dict))
            parsed_llama_dict = parts_analyzer.parse_string(extract_value_llama_dict, reference_prefix)
            parsed_value_llama_dict = {k: v for k, v in parsed_llama_dict.items() if k not in "productName"}
        if not bool(parsed_value_llama_dict):
            original_classification.update(parsed_value_llama_dict)


def search_pcb_column_cols_pandas(df, start_item_index, header_column_search_list):
    """
    pandas DataFrame을 사용하여 PCB 열 열을 검색합니다.

    매개변수:
        df (pandas.DataFrame): 검색할 DataFrame입니다.
        start_item_index (int): 시작 행의 인덱스입니다.
        header_column_search_list (list): 대상 열에 대한 정보를 포함하는 딕셔너리의 리스트입니다.

    반환값:
        list: 각 열에 대한 기계 학습 결과의 리스트입니다.
    """
    ml_results = []

    # DataFrame의 각 열을 반복합니다.
    for col_idx, col_name in enumerate(df.columns):
        col = df[col_name]
        col_list = []
        ml_result = {}

        # start_item_index 이후의 행들을 반복합니다.
        for row_idx in range(start_item_index, len(col)):
            val = col[row_idx]
            column_target = None
            if len(header_column_search_list) > row_idx:
                column_target = header_column_search_list[row_idx].get('target', None)

            col_list.append({
                'type': type(val).__name__,
                'rowIdx': row_idx,
                'columnTarget': column_target,
                'query': str(val).strip() if val is not None else ''
            })

        # 머신러닝 모델이나 분석 로직을 적용합니다.
        if col_list:
            col_query = [l['query'] for l in col_list]
            if direct_analysis.percent_reference(col_query, 55) >= 55:
                ml_results.append({'columnIdx': col_idx, 'target': REFERENCE, 'averageScore': 100, 'isPcbItem': True})
            elif direct_analysis.percent_classification(col_query) >= 40:
                ml_results.append({'columnIdx': col_idx, 'target': VALUE, 'averageScore': 100, 'isPcbItem': True})
            elif direct_analysis.percent_package(col_query) >= 60:
                ml_results.append({'columnIdx': col_idx, 'target': PACKAGE, 'averageScore': 100, 'isPcbItem': True})
            elif direct_analysis.percent_increment_digit_by_list(col_query) >= 70:
                ml_results.append(
                    {'columnIdx': col_idx, 'target': NO, 'averageScore': 100, 'isPcbItem': False})  # 99 는 No
            elif direct_analysis.percent_digit_by_list(col_query) >= 75:
                ml_results.append(
                    {'columnIdx': col_idx, 'target': QTY, 'averageScore': 100, 'isPcbItem': False})  # 4 는 수량
            elif direct_analysis.starts_with_by_list(col_query, '=') >= 75:
                ml_results.append(
                    {'columnIdx': col_idx, 'target': EXCEL_CAL, 'averageScore': 100, 'isPcbItem': False})  # 100 는 엑셀서식
            elif direct_analysis.percent_valid_urls(col_query) >= 75:
                ml_results.append(
                    {'columnIdx': col_idx, 'target': URL, 'averageScore': 100, 'isPcbItem': False})  # 98 는 URL
            else:
                find_result = search_service.find_pcb_item(col_list)
                find_result['columnIdx'] = col_idx
                ml_results.append(find_result)

    return ml_results


def hamming_distance(s1, s2):
    # 두 문자열의 길이를 같게 조정
    l1, l2 = len(s1), len(s2)
    if l1 > l2:
        s2 += ' ' * (l1 - l2)  # s2를 s1의 길이에 맞추어 패딩
    elif l2 > l1:
        s1 += ' ' * (l2 - l1)  # s1을 s2의 길이에 맞추어 패딩

    # Hamming Distance 계산
    return sum(ch1 != ch2 for ch1, ch2 in zip(s1, s2))


def similarity(a, b):
    return SequenceMatcher(None, a, b).ratio()


def levenshtein_distance(s1, s2):
    # 두 문자열의 길이를 가져옴
    len_s1, len_s2 = len(s1), len(s2)

    # 두 문자열 길이의 매트릭스를 생성
    dp = [[0 for _ in range(len_s2 + 1)] for _ in range(len_s1 + 1)]

    # 초기값 설정
    for i in range(len_s1 + 1):
        dp[i][0] = i
    for j in range(len_s2 + 1):
        dp[0][j] = j

    # Levenshtein Distance 계산
    for i in range(1, len_s1 + 1):
        for j in range(1, len_s2 + 1):
            cost = 0 if s1[i - 1] == s2[j - 1] else 1
            dp[i][j] = min(dp[i - 1][j] + 1,  # 삭제
                           dp[i][j - 1] + 1,  # 삽입
                           dp[i - 1][j - 1] + cost)  # 교체 또는 유지
    return dp[len_s1][len_s2]


def load_bom_excel_pandas(filename, sheet_idx=0):
    file_path = './upload/' + filename
    fname, ext = os.path.splitext(filename)
    if ext == '.xls':
        # xls 파일을 xlsx로 변환하는 로직 필요
        # cvt_xls_to_xlsx(file_path, file_path + 'x')
        file_path = file_path + 'x'

    # Pandas를 사용하여 Excel 파일 읽기
    df = pd.read_excel(file_path, sheet_name=sheet_idx, header=None)

    # DataFrame의 각 열에 대해 float 데이터 유형을 object로 변환하고, NaN 및 빈 문자열을 None으로 변환
    for col in df.columns:
        if df[col].dtype == float:
            df[col] = df[col].astype(object)
        df[col] = df[col].where(pd.notna(df[col]) & (df[col] != ''), None)

    return df


def update_col_analysis_to_header(header_list, cols_detail):
    for header in header_list:
        for cols in cols_detail:
            if (header['columnIdx'] == cols['columnIdx'] and cols['averageScore'] > 90
                    and (header['glScore'] is None or cols['averageScore'] > header['glScore'])):
                header['target'] = cols['target']
                header['glScore'] = 100
                header['averageScore'] = 100


def indexing_pcb_item(target, item_name):
    url = "http://localhost:8080/api/pcbItem/_indexing"
    headers = {
        'Content-Type': 'application/x-www-form-urlencoded; charset=UTF-8'
    }
    data = {
        'target': target,
        'itemName': item_name
    }
    response = reqs.post(url, headers=headers, data=data)

    if response.status_code == 200:
        return response.json()
    else:
        return f"Error: {response.status_code}"


def analyzer_file(filename):
    df = load_bom_excel_pandas(filename)

    header_result = search_pcb_header_each_pandas(df)
    header_idx = header_result['headerColumnIdx'] + 1
    header_list = header_result['headerDetail']['pcbColumnSearchList']
    cols_detail = search_pcb_column_cols_pandas(df, header_idx, header_list)
    update_col_analysis_to_header(header_list, cols_detail)
    item_detail = search_pcb_column_each_pandas(df, header_idx, header_list)

    data = {
        'headerColumnIdx': header_idx,
        'headerList': header_list,
        'colsDetail': cols_detail,
        'itemDetail': item_detail
    }
    return {'result': True, 'data': data}
