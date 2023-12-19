import collections
import json
import numbers
import os
import re
from urllib.parse import urlparse

import numpy as np
import openpyxl
import pandas as pd
import requests as reqs
import xlrd
from openpyxl.workbook import Workbook
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.linear_model import LogisticRegression
from sklearn.metrics import accuracy_score
from sklearn.model_selection import train_test_split
from data_extractor import DataExtractor
from parts_analysis import PartsAnalysis

extractor = DataExtractor()
analyzer = PartsAnalysis()

tfidf_vect = TfidfVectorizer(stop_words='english', ngram_range=(1, 1), max_df=100)
lr_clf = LogisticRegression(C=5)

REFERENCE = 1
PART_NUMBER = 2
PACKAGE = 6
VALUE = 9


# 엑셀 파일 로드
# 컬럼 카운트 리스트, sheet 반환
def load_bom_excel(filename, sheet_idx=0):
    file_path = './upload/' + filename
    fname, ext = os.path.splitext(filename)
    if ext == '.xls':
        cvt_xls_to_xlsx(file_path, file_path + 'x')
        file_path = file_path + 'x'

    bom_excel = openpyxl.load_workbook(file_path, data_only=True)

    # 행별 실제 사용하는 컬럼 개수 알아내기
    sheet = bom_excel.worksheets[sheet_idx]

    column_cnt_list = {}
    for idx, row in enumerate(sheet.rows):
        none_cnt = 0
        for r in row:
            if r.value is not None:
                none_cnt = none_cnt + 1
            column_cnt_list[idx] = none_cnt

    return column_cnt_list, sheet


# https://stackoverflow.com/questions/9918646/how-to-convert-xls-to-xlsx/42574983#42574983
def cvt_xls_to_xlsx(src_file_path, dst_file_path):
    book_xls = xlrd.open_workbook(src_file_path)
    book_xlsx = Workbook()

    sheet_names = book_xls.sheet_names()
    for sheet_index, sheet_name in enumerate(sheet_names):
        sheet_xls = book_xls.sheet_by_name(sheet_name)
        if sheet_index == 0:
            sheet_xlsx = book_xlsx.active
            sheet_xlsx.title = sheet_name
        else:
            sheet_xlsx = book_xlsx.create_sheet(title=sheet_name)

        for row in range(0, sheet_xls.nrows):
            for col in range(0, sheet_xls.ncols):
                sheet_xlsx.cell(row=row + 1, column=col + 1).value = sheet_xls.cell_value(row, col)

    book_xlsx.save(dst_file_path)


# 가장 큰 값 알아내기
def get_max_count(tuple_data):
    data_count = collections.Counter(val for val in tuple_data.values())

    def f1(x):
        return data_count[x]

    max_cnt = max(data_count.keys(), key=f1)
    return max_cnt


# tuple 데이터로부터 최대값을 가진 인덱스를 반환
def get_index_max_value(tuple_data):
    def f1(x):
        return tuple_data[x]

    return max(tuple_data, key=f1)


def get_header_column_idx(data, max_column_cnt):
    header_column_idx = -1
    for d in data:
        if data[d] == max_column_cnt:
            header_column_idx = d
            break

    return header_column_idx


def search_pcb_column(sheet, header_column_idx):
    result = []
    for row in tuple(sheet.rows)[header_column_idx]:
        if row.value is None:
            continue

        URL = 'http://localhost:8080/api/pcbColumn/_searchSentence?q={0}'.format(row.value)
        response = reqs.get(URL)
        # response.status_code
        body = json.loads(response.text)
        if body['result'] != True:
            continue

        if not body['data']:
            continue
        # print(body['data'][0]['target'])
        result.append({'importColName': row.value, 'searchInfo': body['data'][0]})
    return result


def search_excel_column(query):
    data, sheet = load_bom_excel(query)
    max_column_cnt = get_max_count(data)
    header_column_idx = get_header_column_idx(data, max_column_cnt)
    ccResult = {'result': True, 'data': {'headerColumnIdx': header_column_idx + 1, 'maxColumnCnt': max_column_cnt,
                                         'columnInfo': search_pcb_column(sheet, header_column_idx)}}
    return ccResult


# def search_pcb_header_by_row(sheet, header_column_idx):
#     result = []
#     row_list = []
#     for row in tuple(sheet.rows)[header_column_idx]:
#         if row.value is None:
#             continue
#         row_list.append(row.value)
#
#     URL = 'http://localhost:8080/api/pcbColumn/_searchSentenceList'
#     print(row_list)
#     reqs.post(URL, json={'queryColumnNameList': row_list})


def search_pcb_header_each(sheet):
    sentence_results = []
    score_tuple = {}
    score_tuple_for_bom = {}
    for (idx, rows) in enumerate(sheet.rows):
        row_list = []
        row_list_for_bom = []
        for row in rows:
            if row.value is not None:
                row_list.append(str(row.value))
            else:
                row_list.append('')

            column_target = None
            if row.value is None:
                column_target = None
            row_list_for_bom.append({'type': type(row.value).__name__, 'columnIdx': idx, 'columnTarget': column_target,
                                     'query': str(row.value).strip()})

        URL = 'http://localhost:8080/api/pcbColumn/_searchSentenceList'
        response = reqs.post(URL, json={'queryColumnNameList': row_list})
        body = json.loads(response.text)
        response_data = body['data']
        score_tuple[idx] = response_data['averageScore']
        sentence_results.append(response_data)
        score_tuple_for_bom[idx] = bom_ml(row_list_for_bom, True)
        if idx == 13:
            break

    header_column_idx = get_index_max_value(score_tuple)
    if score_tuple_for_bom[header_column_idx]['averageScore'] < 50:
        header_column_idx = header_column_idx + 1
    return {'headerColumnIdx': header_column_idx + 1, 'headerDetail': sentence_results[header_column_idx]}


def bom_ml_test():
    samplepcb_df = pd.read_csv('./data/samplepcb_bom.csv', header=None)

    X = samplepcb_df.iloc[:, :-1]
    y = samplepcb_df.iloc[:, -1]

    X_train, X_test, y_train, y_test = train_test_split(X, y, test_size=0.2, random_state=156, stratify=y)

    X_train = X_train[0].tolist()
    X_test = X_test[0].tolist()
    tfidf_vect = TfidfVectorizer(stop_words='english', ngram_range=(1, 1), max_df=100)
    tfidf_vect.fit(X_train)
    X_train_tfidf_vect = tfidf_vect.transform(X_train)
    X_test_tfidf_vect = tfidf_vect.transform(X_test)

    lr_clf = LogisticRegression(C=5)
    lr_clf.fit(X_train_tfidf_vect, y_train)
    pred = lr_clf.predict(X_test_tfidf_vect)
    print('TF-IDF Vectorized Logistic Regression 의 예측 정확도는 {0:.3f}'.format(accuracy_score(y_test, pred)))


def bom_ml_tuning():
    samplepcb_df = pd.read_csv('./data/samplepcb_bom.csv', header=None)

    X = samplepcb_df.iloc[:, :-1]
    y = samplepcb_df.iloc[:, -1]

    X_train, X_test, y_train, y_test = train_test_split(X, y, test_size=0.2, random_state=156, stratify=y)

    X_train = X_train[0].tolist()
    X_test = X_test[0].tolist()

    from sklearn.model_selection import GridSearchCV
    from sklearn.pipeline import Pipeline

    pipeline = Pipeline([
        ('tfidf_vect', TfidfVectorizer(stop_words='english')),
        ('lr_clf', LogisticRegression())
    ])

    # Pipeline에 기술된 각각의 객체 변수에 언더바(_)2개를 연달아 붙여 GridSearchCV에 사용될
    # 파라미터/하이퍼 파라미터 이름과 값을 설정. .
    params = {'tfidf_vect__ngram_range': [(1, 1), (1, 2), (1, 3)],
              'tfidf_vect__max_df': [100, 300, 700],
              'lr_clf__C': [1, 5, 10]
              }

    # GridSearchCV의 생성자에 Estimator가 아닌 Pipeline 객체 입력
    grid_cv_pipe = GridSearchCV(pipeline, param_grid=params, cv=3, scoring='accuracy', verbose=1)
    grid_cv_pipe.fit(X_train, y_train)
    print(grid_cv_pipe.best_params_, grid_cv_pipe.best_score_)

    pred = grid_cv_pipe.predict(X_test)
    print('Pipeline을 통한 Logistic Regression 의 예측 정확도는 {0:.3f}'.format(accuracy_score(y_test, pred)))


def bom_ml_init():
    global tfidf_vect
    global lr_clf

    samplepcb_df = pd.concat(pd.read_excel('./data/samplepcb_bom.xlsx', header=None, sheet_name=None),
                             ignore_index=True)

    X = samplepcb_df.iloc[:, :-1]
    y = samplepcb_df.iloc[:, -1]

    X_train = [str(x) for x in X[0].tolist()]
    y_train = y.tolist()

    tfidf_vect.fit(X_train)
    X_train_tfidf_vect = tfidf_vect.transform(X_train)

    lr_clf.fit(X_train_tfidf_vect, y_train)


def bom_ml_init_by_api():
    global tfidf_vect
    global lr_clf

    tfidf_vect = TfidfVectorizer(stop_words='english', ngram_range=(1, 1), max_df=100)
    lr_clf = LogisticRegression(C=5)

    URL = 'http://localhost:8080/api/pcbItem/_allItemGroupByTarget'
    response = reqs.get(URL)
    body = json.loads(response.text)

    X_train = []
    y_train = []
    for items in body['data']:
        for item in items:
            X_train.append(item['itemName'])
            y_train.append(item['target'])

    tfidf_vect.fit(X_train)
    X_train_tfidf_vect = tfidf_vect.transform(X_train)

    lr_clf.fit(X_train_tfidf_vect, y_train)


def bom_ml(query_list, zero_score=False):
    global tfidf_vect
    global lr_clf

    query_values = [l['query'] for l in query_list]
    query_tfidf_vect = tfidf_vect.transform(query_values)
    predict_list = lr_clf.predict(query_tfidf_vect)
    predict_proba_list = lr_clf.predict_proba(query_tfidf_vect)
    score_list = []
    results = []
    for idx, meta in enumerate(query_list):
        score = max(predict_proba_list[idx]) * 100
        if type(None).__name__ == meta['type']:
            score = 0
            if zero_score:
                score_list.append(score)
        else:
            score_list.append(score)

        meta['predict'] = predict_list[idx].item()
        meta['score'] = score
        # classification 결과를 받아옴
        classification_result = {k: v for k, v in classification(meta['query']).items() if k != 'productName'}
        # classification_result가 비어 있지 않은 경우에만 meta['classification']에 값을 넣음
        if classification_result:
            meta['classification'] = classification_result
        results.append(meta)

    if len(score_list) == 0:
        score_list.append(0)

    return {'predictResults': results, 'averageScore': np.mean(score_list)}


def is_increment_digit_by_list(any_list, percent):
    """
    리스트가 전부 숫자인지와 증가하는지 체크
    :param any_list: 리스트
    :return:
    """
    is_digit = False
    pre_val = False
    increment_cnt = 0
    none_digit_cnt = 0
    any_list_len = len(any_list)
    for val in any_list:
        if isinstance(val, str) and val.isdigit():  # 문자형이지만 숫자라면
            is_digit = True
        elif isinstance(val, numbers.Number):  # 숫자라면
            is_digit = True
        elif val is None or val == 'None':
            # none_digit_cnt = none_digit_cnt + 1
            any_list_len = any_list_len - 1
            continue
        else:
            continue

        if is_digit:
            val = int(val)

        if pre_val and pre_val < val:
            increment_cnt = increment_cnt + 1

        pre_val = val

    return is_digit and (increment_cnt / (any_list_len - none_digit_cnt)) * 100 >= percent


def is_digit_by_list(any_list, percent):
    """
    리스트가 전부 숫자인지와 증가하는지 체크
    :param any_list: 리스트
    :return:
    """
    any_list_len = len(any_list)
    digit_cnt = 0
    not_digit_cnt = 0

    for val in any_list:
        if isinstance(val, str) and val.isdigit():  # 문자형이지만 숫자라면
            digit_cnt = digit_cnt + 1
        elif isinstance(val, numbers.Number):  # 숫자라면
            digit_cnt = digit_cnt + 1
        elif val is None or val == 'None':
            any_list_len = any_list_len - 1
            continue
        else:
            not_digit_cnt = not_digit_cnt + 1

    if any_list_len == 0:
        return False
    return (digit_cnt / any_list_len) * 100 >= percent


def starts_with_by_list(any_list, str, percent):
    """
    리스트가 전부 숫자인지와 증가하는지 체크
    :param any_list: 리스트
    :return:
    """
    if not any_list:
        return False

    s_cnt = 0
    for val in any_list:
        if val[0:1] == str:
            s_cnt = s_cnt + 1

    return (s_cnt / len(any_list)) * 100 >= percent


def is_valid_url(url):
    parsed_url = urlparse(url)
    return bool(parsed_url.scheme) and bool(parsed_url.netloc)


def percent_valid_urls(url_list, percent):
    if not url_list:
        return False

    valid_urls = sum(is_valid_url(url) for url in url_list)
    valid_percent = (valid_urls / len(url_list)) * 100

    return valid_percent >= percent


def calculate_none_percent(strings):
    none_count = sum(1 for item in strings if item is None or item == "None")
    none_percent = (none_count / len(strings)) * 100 if strings else 0
    return none_percent


def trim_none_or_empty_from_end(strings):
    # 리스트 끝에서부터 None, 'None' 또는 빈 문자열이 아닌 첫 번째 요소 찾기
    first_non_none_or_empty = next((i for i, item in enumerate(reversed(strings)) if item and item != "None"),
                                   len(strings))

    # 해당 인덱스를 사용하여 필요한 부분만 남기기
    return strings[:len(strings) - first_non_none_or_empty]


def percent_reference(item_list, percent, none_exclude=True):
    if not item_list:
        return False

    pattern = re.compile(r"(?:^|,)[CRUJLSD]\d+(?:,|$)")

    none_percent = calculate_none_percent(item_list)

    # None 비율이 주어진 퍼센트 이상인 경우 False 반환
    if none_percent >= percent:
        return False

    if none_exclude:
        item_list = [item for item in item_list if item is not None and item != "None"]

    matching = sum(bool(pattern.match(string)) for string in item_list)
    matching_percent = (matching / len(item_list)) * 100

    return matching_percent >= percent


def classification(item_val):
    title_text = analyzer.split_text(item_val)
    return analyzer.parse_string(title_text)


def percent_classification(item_list, percent, none_exclude=True):
    successful_classifications = 0

    if none_exclude:
        item_list = [item for item in item_list if item is not None and item != "None"]

    # item_list가 비어있거나 모든 항목이 제거된 경우 0을 반환하도록 처리
    if not item_list:
        return False

    for item_val in item_list:
        title_text = analyzer.split_text(item_val)
        classifications = analyzer.parse_string(title_text)

        key_count = len([key for key in classifications if key != 'productName'])

        if key_count >= 1:
            successful_classifications += 1

    success_percent = (successful_classifications / len(item_list)) * 100
    return success_percent >= percent


def percent_package(str_list, percent):
    if not str_list:
        return False

    successful_package = 0
    for item_val in str_list:
        size = extractor.extract_size_from_title(item_val)
        if size:
            successful_package += 1

    # 성공 비율 계산
    success_percent = (successful_package / len(str_list)) * 100
    return success_percent >= percent


def percent_none_for_is_pcb(item_list, percent):
    none_count = 0
    # 숫자가 아닌 요소만 필터링 (문자열 형태의 숫자도 포함)
    filtered_list = [item for item in item_list if
                     not (isinstance(item, (int, float)) or (isinstance(item, str) and item.isdigit()))]

    for item_val in filtered_list:
        title_text = analyzer.split_text(item_val)
        classifications = analyzer.parse_string(title_text)

        # '품명' 키를 제외하고 키의 개수를 계산
        key_count = len([key for key in classifications if key != 'productName'])

        if key_count >= 1:
            return False

        # None 값이거나 None 타입인 경우 카운트
        if item_val is None or item_val == "None":
            none_count += 1

    # None 값의 비율 계산
    none_percent = (none_count / len(filtered_list)) * 100 if filtered_list else 0
    return none_percent >= percent


def find_update_part_number_average_score(results):
    # POST 요청을 보낼 URL
    url = 'http://localhost:8080/api/pcbItem/_searchList?target=2'

    for result in results:
        if result['predict'] == 2:
            item = result['predictResults']
            # POST 요청을 보내고 응답을 받음
            response = reqs.post(url, json=[l['query'] for l in item])

            if response.status_code == 200:
                data = response.json().get('data', [])

                # _score 값들을 추출
                scores = [record['_score'] for record in data]

                # 평균 점수 계산
                if scores:
                    average_score = sum(scores) / len(scores)
                else:
                    average_score = 0  # 점수가 없는 경우 평균 점수를 0으로 설정

                # 해당 항목에 averageScore 추가
                result['averageScore'] = average_score


def bom_ml_cols(query_list):
    global tfidf_vect
    global lr_clf

    query_values = [l['query'] for l in query_list]
    query_tfidf_vect = tfidf_vect.transform(query_values)
    predict_list = lr_clf.predict(query_tfidf_vect)
    predict_proba_list = lr_clf.predict_proba(query_tfidf_vect)
    score_list = []
    results = []
    for idx, meta in enumerate(query_list):
        score = max(predict_proba_list[idx]) * 100
        if type(None).__name__ == meta['type']:
            score = 0
            # score_list.append(score) # None, 0 일때는 평균점수에 반영하지 않는다
        if meta['query'] == 'None':
            score = 0
        else:
            score_list.append(score)

        if len(score_list) == 0:
            score_list.append(0)

        meta['predict'] = predict_list[idx].item()
        meta['score'] = score
        results.append(meta)

    queries = [result['query'] for result in results]
    # 리스트 끝에서부터 None, 'None' 또는 빈 문자열이 아닌 첫 번째 요소 찾기
    queries = trim_none_or_empty_from_end(queries)
    if percent_reference(queries, 70):
        return {'predict': REFERENCE, 'predictResults': results, 'averageScore': 100}
    if percent_classification(queries, 40):
        return {'predict': VALUE, 'predictResults': results, 'averageScore': 100}
    if percent_package(queries, 60):
        return {'predict': PACKAGE, 'predictResults': results, 'averageScore': 100}
    if is_increment_digit_by_list(queries, 80):
        return {'predict': 99, 'predictResults': results, 'averageScore': 100}  # 99 는 No
    if is_digit_by_list(queries, 75):
        return {'predict': 4, 'predictResults': results, 'averageScore': 100}  # 4 는 수량
    if starts_with_by_list(queries, '=', 75):
        return {'predict': 100, 'predictResults': results, 'averageScore': 100}  # 100 는 엑셀서식
    if percent_valid_urls(queries, 75):
        return {'predict': 98, 'predictResults': results, 'averageScore': 100}  # 98 는 URL

    predicts = [result['predict'] for result in results]

    data_count = collections.Counter(predicts)

    def f1(x):
        return data_count[x]

    max_cnt = max(data_count.keys(), key=f1)

    # 예상된 분류의 점수값만 추가하여 평균계산
    predict_score_list = []
    none_score = 0
    for result in results:
        if result['predict'] == max_cnt and result['score'] != 0:
            predict_score_list.append(result['score'])
        if result['score'] == 0:
            none_score = none_score + 1

    if len(predict_score_list) == 0:
        predict_score_list.append(0)

    minus_score = (none_score / len(results)) * 100
    return {'predict': max_cnt, 'predictResults': results, 'averageScore': np.mean(predict_score_list) - minus_score}


def search_pcb_column_each(sheet, start_item_index, header_column_search_list):
    """
PCB 컬럼 분석검색
    :param sheet: 시트
    :param start_item_index: item 시작 인덱스
    :param header_column_search_list: 분석된 헤더컬럼 정보
    :return:
    """
    ml_results = []
    column_idx = start_item_index
    for col_idx, rows in enumerate(sheet.iter_rows(min_row=start_item_index)):
        row_list = []
        ml_result = {}
        for idx, row in enumerate(rows):
            column_target = None
            if len(header_column_search_list) > idx:
                column_target = header_column_search_list[idx]['target']
            if row.value is None:
                column_target = None
            row_list.append({'type': type(row.value).__name__, 'columnIdx': idx, 'columnTarget': column_target,
                             'query': str(row.value).strip()})

        is_pcb_item = 'isPcbItem'
        if len(row_list) != 0:
            if percent_none_for_is_pcb([l['query'] for l in row_list], 90):
                ml_result[is_pcb_item] = False
            else:
                ml_result = bom_ml(row_list)
                if ml_result['averageScore'] > 50:
                    ml_result[is_pcb_item] = True
                else:
                    ml_result[is_pcb_item] = False
        else:
            ml_result[is_pcb_item] = False

        ml_result['columnIdx'] = column_idx
        column_idx = column_idx + 1
        ml_results.append(ml_result)

    return ml_results


def search_pcb_column_cols(sheet, start_item_index, header_column_search_list):
    ml_results = []
    for col_idx, cols in enumerate(sheet.iter_cols(min_row=start_item_index)):
        col_list = []
        ml_result = {}
        for row_idx, col in enumerate(cols):
            column_target = None
            if len(header_column_search_list) > row_idx:
                column_target = header_column_search_list[col_idx]['target']
            if col.value is None:
                column_target = None
            col_list.append({'type': type(col.value).__name__, 'rowIdx': row_idx, 'columnTarget': column_target,
                             'query': str(col.value).strip()})

        if len(col_list) != 0:
            ml_result = bom_ml_cols(col_list)

        ml_results.append(ml_result)

    find_update_part_number_average_score(ml_results)
    return ml_results


def analysis_bom(query):
    """
    BOM 분석
    :param query: 파일명
    :return:
    """
    column_cnt_list, sheet = load_bom_excel(query)
    search_result = search_pcb_header_each(sheet)
    item_detail = search_pcb_column_each(sheet, search_result['headerColumnIdx'] + 1,
                                         search_result['headerDetail']['pcbColumnSearchList'])
    cols_detail = search_pcb_column_cols(sheet, search_result['headerColumnIdx'] + 1,
                                         search_result['headerDetail']['pcbColumnSearchList'])
    max_column_cnt = get_max_count(column_cnt_list)

    data = {
        'headerColumnIdx': search_result['headerColumnIdx'],
        'maxColumnCnt': max_column_cnt,
        'headerDetail': search_result['headerDetail'],
        'itemDetail': item_detail,
        'colsDetail': cols_detail
    }
    ccResult = {'result': True, 'data': data}
    return ccResult
